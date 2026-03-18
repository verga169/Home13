import io
import json
import os
import re
import sys
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timedelta
import unicodedata
from uuid import uuid4

from flask import Flask, Response, jsonify, redirect, render_template, request, send_from_directory, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash

try:
    import psycopg
    from psycopg.rows import dict_row
except Exception:
    psycopg = None
    dict_row = None

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, "data_store.json")


def load_local_env() -> None:
    """Load basic KEY=VALUE pairs from .env/.env.local if env vars are missing."""
    for filename in (".env.local", ".env"):
        env_path = os.path.join(BASE_DIR, filename)
        if not os.path.exists(env_path):
            continue

        with open(env_path, "r", encoding="utf-8") as env_file:
            for raw_line in env_file:
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value


load_local_env()
DATABASE_URL = (os.environ.get("DATABASE_URL") or "").strip()
USE_DATABASE = bool(DATABASE_URL)
GEMINI_MODEL = (os.environ.get("GEMINI_MODEL") or "gemini-2.5-flash-lite").strip()

AUTH_USERNAME = (os.environ.get("HOME13_AUTH_USERNAME") or "admin").strip()
AUTH_PASSWORD = os.environ.get("HOME13_AUTH_PASSWORD")
AUTH_PASSWORD_HASH = (os.environ.get("HOME13_AUTH_PASSWORD_HASH") or "").strip()

ROLE_SUPERADMIN = "superadmin"
ROLE_USER = "user"

IS_RENDER_DEPLOY = bool((os.environ.get("RENDER") or "").strip() or (os.environ.get("RENDER_EXTERNAL_URL") or "").strip())
LOCAL_SUPERADMIN_USERNAME = "admin"

app.secret_key = (
    (os.environ.get("FLASK_SECRET_KEY") or "").strip()
    or (os.environ.get("SECRET_KEY") or "").strip()
    or "home13-dev-secret-change-me"
)

SESSION_DAYS_RAW = (os.environ.get("HOME13_SESSION_DAYS") or "90").strip()
try:
    SESSION_DAYS = max(1, int(SESSION_DAYS_RAW))
except ValueError:
    SESSION_DAYS = 90


def _read_env_int(var_name: str, default_value: int, min_value: int, max_value: int) -> int:
    raw_value = (os.environ.get(var_name) or "").strip()
    if not raw_value:
        return default_value
    try:
        return max(min_value, min(max_value, int(raw_value)))
    except ValueError:
        return default_value


AI_MAX_HISTORY_TURNS = _read_env_int("HOME13_AI_MAX_HISTORY_TURNS", 12, 2, 80)
AI_MAX_OUTPUT_TOKENS = _read_env_int("HOME13_AI_MAX_OUTPUT_TOKENS", 380, 64, 2048)
AI_MAX_ITERATIONS = _read_env_int("HOME13_AI_MAX_ITERATIONS", 4, 1, 10)

app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=SESSION_DAYS)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = bool(
    (os.environ.get("SESSION_COOKIE_SECURE") or "").strip() == "1" or IS_RENDER_DEPLOY
)

DEFAULT_DATA = {
    "expenses": {
        "acquisto_casa": [],
        "ristrutturazione": [],
    },
    "loans": [],
    "repayments": [],
}

CATEGORY_DISPLAY_NAMES = {
    "acquisto_casa": "Acquisto Casa",
    "ristrutturazione": "Ristrutturazione",
    "loans": "Prestiti",
    "repayments": "Rimborsi",
}

SECTION_ALIASES = {
    "acquisto_casa": "acquisto_casa",
    "acquisto casa": "acquisto_casa",
    "acquisto": "acquisto_casa",
    "ristrutturazione": "ristrutturazione",
    "ristrutturazioni": "ristrutturazione",
    "loans": "loans",
    "prestiti": "loans",
    "prestiti ricevuti": "loans",
    "repayments": "repayments",
    "rimborsi": "repayments",
    "all": "all",
    "tutto": "all",
    "tutti": "all",
}


def get_gemini_api_key() -> str:
    # Re-read local env files so updates to .env.local are picked up immediately.
    load_local_env()
    return (os.environ.get("GEMINI_API_KEY") or "").strip()


def get_gemini_model() -> str:
    load_local_env()
    return (os.environ.get("GEMINI_MODEL") or GEMINI_MODEL or "gemini-2.5-flash-lite").strip()


def is_authenticated() -> bool:
    return bool(session.get("authenticated"))


def is_superadmin_session() -> bool:
    return session.get("role") == ROLE_SUPERADMIN


def normalize_username(raw_value: str) -> str:
    return sanitize_text(raw_value).casefold()


def get_superadmin_username() -> str:
    if IS_RENDER_DEPLOY:
        return sanitize_text(AUTH_USERNAME)
    return LOCAL_SUPERADMIN_USERNAME


def _is_safe_next_path(next_path: str) -> bool:
    target = sanitize_text(next_path)
    return bool(target) and target.startswith("/") and not target.startswith("//")


def verify_superadmin_credentials(username: str, password: str) -> bool:
    superadmin_username = get_superadmin_username()
    if normalize_username(username) != normalize_username(superadmin_username):
        return False

    raw_password = password or ""

    if AUTH_PASSWORD_HASH:
        try:
            return check_password_hash(AUTH_PASSWORD_HASH, raw_password)
        except Exception:
            return False

    if AUTH_PASSWORD is not None:
        return raw_password == AUTH_PASSWORD

    # In local mode allow bootstrap admin/admin. On Render require env credentials.
    if IS_RENDER_DEPLOY:
        return False
    return raw_password == "admin"


def fetch_db_user_by_username(username: str) -> dict | None:
    if not USE_DATABASE:
        return None

    candidate = sanitize_text(username)
    if not candidate:
        return None

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT username, password_hash, is_active
                FROM users
                WHERE LOWER(username) = LOWER(%s)
                """,
                (candidate,),
            )
            return cur.fetchone()


def authenticate_login(username: str, password: str) -> dict | None:
    candidate_username = sanitize_text(username)
    raw_password = password or ""
    if not candidate_username:
        return None

    superadmin_username = get_superadmin_username()
    if normalize_username(candidate_username) == normalize_username(superadmin_username):
        if verify_superadmin_credentials(candidate_username, raw_password):
            return {"username": superadmin_username, "role": ROLE_SUPERADMIN}
        return None

    if not USE_DATABASE:
        return None

    try:
        db_user = fetch_db_user_by_username(candidate_username)
    except Exception:
        return None

    if not db_user or not bool(db_user.get("is_active")):
        return None

    password_hash = sanitize_text(db_user.get("password_hash"))
    if not password_hash:
        return None

    try:
        if not check_password_hash(password_hash, raw_password):
            return None
    except Exception:
        return None

    return {"username": sanitize_text(db_user.get("username")) or candidate_username, "role": ROLE_USER}


def format_euro(value: float) -> str:
    formatted = f"{float(value):,.2f}"
    integer, decimal = formatted.split(".")
    integer = integer.replace(",", ".")
    return f"{integer},{decimal}"


app.jinja_env.filters["euro"] = format_euro


def format_date_it(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    try:
        parsed = datetime.fromisoformat(raw[:10])
        return parsed.strftime("%d/%m/%Y")
    except ValueError:
        return raw


app.jinja_env.filters["date_it"] = format_date_it


def _asset_mtime(relative_path: str) -> str:
    file_path = os.path.join(BASE_DIR, relative_path)
    try:
        return str(int(os.path.getmtime(file_path)))
    except OSError:
        return "1"


@app.context_processor
def inject_asset_urls() -> dict:
    favicon_v = _asset_mtime(os.path.join("static", "favicon.ico"))
    favicon16_v = _asset_mtime(os.path.join("static", "favicon-16x16.png"))
    favicon32_v = _asset_mtime(os.path.join("static", "favicon-32x32.png"))
    android192_v = _asset_mtime(os.path.join("static", "android-chrome-192x192.png"))
    android512_v = _asset_mtime(os.path.join("static", "android-chrome-512x512.png"))
    apple_touch_v = _asset_mtime(os.path.join("static", "apple-touch-icon.png"))
    manifest_v = _asset_mtime(os.path.join("static", "site.webmanifest"))
    sw_v = _asset_mtime(os.path.join("static", "sw.js"))
    return {
        "favicon_ico_url": url_for("static", filename="favicon.ico", v=favicon_v),
        "favicon_16_url": url_for("static", filename="favicon-16x16.png", v=favicon16_v),
        "favicon_32_url": url_for("static", filename="favicon-32x32.png", v=favicon32_v),
        "android_192_url": url_for("static", filename="android-chrome-192x192.png", v=android192_v),
        "android_512_url": url_for("static", filename="android-chrome-512x512.png", v=android512_v),
        "apple_touch_icon_url": url_for("static", filename="apple-touch-icon.png", v=apple_touch_v),
        "manifest_url": url_for("static", filename="site.webmanifest", v=manifest_v),
        "sw_url": url_for("service_worker", v=sw_v),
    }


def parse_amount(raw_value: str) -> float:
    raw = (raw_value or "").strip().replace(" ", "")
    if not raw:
        raise ValueError("Importo mancante")

    has_dot = "." in raw
    has_comma = "," in raw

    if has_dot and has_comma:
        # Italian style: 1.234,56
        normalized = raw.replace(".", "").replace(",", ".")
    elif has_comma:
        # Decimal comma: 1234,56
        normalized = raw.replace(",", ".")
    elif re.fullmatch(r"\d{1,3}(?:\.\d{3})+", raw):
        # Thousands only with dots: 1.234.567
        normalized = raw.replace(".", "")
    else:
        # Plain or dot-decimal input: 1234 or 1234.56
        normalized = raw

    value = float(normalized)
    if value <= 0:
        raise ValueError("Importo deve essere maggiore di zero")
    return value


def sanitize_text(raw_value: str) -> str:
    return (raw_value or "").strip()


def capitalize_first(raw_value: str) -> str:
    """Return the string with only the very first character uppercased."""
    text = sanitize_text(raw_value)
    if not text:
        return text
    return text[0].upper() + text[1:]


def normalize_text(raw_value: str) -> str:
    text = sanitize_text(raw_value).lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip()


def canonical_section(raw_value: str, allow_all: bool = True) -> str:
    normalized = normalize_text(raw_value).replace("_", " ")
    canonical = SECTION_ALIASES.get(normalized, "")
    if canonical == "all" and not allow_all:
        return ""
    return canonical


def parse_iso_date(raw_value: str) -> date:
    raw = sanitize_text(raw_value)
    if not raw:
        return date.today()
    try:
        return datetime.fromisoformat(raw[:10]).date()
    except ValueError as exc:
        raise ValueError("Data non valida") from exc


def delete_ai_operation(intent: str, item_id: str) -> bool:
    if intent == "delete_repayment":
        section = "repayments"
    elif intent == "delete_loan":
        section = "loans"
    elif intent == "delete_expense_ristrutturazione":
        section = "ristrutturazione"
    else:
        section = "acquisto_casa"

    if USE_DATABASE:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                if section in {"acquisto_casa", "ristrutturazione"}:
                    cur.execute("DELETE FROM expenses WHERE id = %s AND category = %s", (item_id, section))
                elif section == "loans":
                    cur.execute("DELETE FROM loans WHERE id = %s", (item_id,))
                else:
                    cur.execute("DELETE FROM repayments WHERE id = %s", (item_id,))
                deleted = cur.rowcount > 0
            if deleted:
                conn.commit()
        return deleted

    data = load_data()
    if section in {"acquisto_casa", "ristrutturazione"}:
        deleted = remove_by_id(data["expenses"][section], item_id)
    elif section == "loans":
        deleted = remove_by_id(data["loans"], item_id)
    else:
        deleted = remove_by_id(data["repayments"], item_id)
    if deleted:
        save_data(data)
    return deleted


def get_db_connection():
    if not USE_DATABASE:
        raise RuntimeError("DATABASE_URL non configurata")
    if psycopg is None or dict_row is None:
        raise RuntimeError("Dipendenza psycopg non installata")
    return psycopg.connect(DATABASE_URL, row_factory=dict_row)


def disable_database_mode(reason: str) -> None:
    """Fall back to local JSON storage when DB is unavailable."""
    global USE_DATABASE, DATABASE_URL
    USE_DATABASE = False
    DATABASE_URL = ""
    print(f"[Home13] Database disabled: {reason}", file=sys.stderr, flush=True)


def ensure_database_ready() -> None:
    if not USE_DATABASE:
        return
    if psycopg is None:
        disable_database_mode(
            "DATABASE_URL presente ma psycopg non disponibile nell'interprete corrente"
        )
        return

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS expenses (
                        id TEXT PRIMARY KEY,
                        category TEXT NOT NULL CHECK (category IN ('acquisto_casa', 'ristrutturazione')),
                        operation_date DATE NOT NULL,
                        description TEXT NOT NULL,
                        amount NUMERIC(14, 2) NOT NULL CHECK (amount > 0)
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS loans (
                        id TEXT PRIMARY KEY,
                        operation_date DATE NOT NULL,
                        lender TEXT NOT NULL,
                        note TEXT NOT NULL DEFAULT '',
                        amount NUMERIC(14, 2) NOT NULL CHECK (amount > 0)
                    )
                    """
                )
                cur.execute("ALTER TABLE loans ADD COLUMN IF NOT EXISTS note TEXT NOT NULL DEFAULT ''")
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS repayments (
                        id TEXT PRIMARY KEY,
                        operation_date DATE NOT NULL,
                        lender TEXT NOT NULL,
                        amount NUMERIC(14, 2) NOT NULL CHECK (amount > 0)
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS users (
                        username TEXT PRIMARY KEY,
                        password_hash TEXT NOT NULL,
                        is_active BOOLEAN NOT NULL DEFAULT TRUE,
                        created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                    )
                    """
                )
            conn.commit()
    except Exception as exc:
        disable_database_mode(f"inizializzazione DB fallita ({exc.__class__.__name__}: {exc})")

def load_data() -> dict:
    if USE_DATABASE:
        try:
            data = json.loads(json.dumps(DEFAULT_DATA))
            with get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT id, category, operation_date, description, amount::DOUBLE PRECISION AS amount
                        FROM expenses
                        """
                    )
                    for row in cur.fetchall():
                        data["expenses"][row["category"]].append(
                            {
                                "id": row["id"],
                                "date": row["operation_date"].isoformat(),
                                "description": row["description"],
                                "amount": float(row["amount"]),
                            }
                        )

                    cur.execute(
                        """
                        SELECT id, operation_date, lender, note, amount::DOUBLE PRECISION AS amount
                        FROM loans
                        """
                    )
                    for row in cur.fetchall():
                        data["loans"].append(
                            {
                                "id": row["id"],
                                "date": row["operation_date"].isoformat(),
                                "lender": row["lender"],
                                "note": sanitize_text(row.get("note")),
                                "amount": float(row["amount"]),
                            }
                        )

                    cur.execute(
                        """
                        SELECT id, operation_date, lender, amount::DOUBLE PRECISION AS amount
                        FROM repayments
                        """
                    )
                    for row in cur.fetchall():
                        data["repayments"].append(
                            {
                                "id": row["id"],
                                "date": row["operation_date"].isoformat(),
                                "lender": row["lender"],
                                "amount": float(row["amount"]),
                            }
                        )

            return data
        except Exception as exc:
            disable_database_mode(f"lettura DB fallita ({exc.__class__.__name__}: {exc})")

    if not os.path.exists(DATA_FILE):
        save_data(DEFAULT_DATA)
        return json.loads(json.dumps(DEFAULT_DATA))

    with open(DATA_FILE, "r", encoding="utf-8") as f:
        loaded = json.load(f)

    data = json.loads(json.dumps(DEFAULT_DATA))
    data["expenses"]["acquisto_casa"] = loaded.get("expenses", {}).get("acquisto_casa", [])
    data["expenses"]["ristrutturazione"] = loaded.get("expenses", {}).get("ristrutturazione", [])
    data["loans"] = loaded.get("loans", [])
    for loan in data["loans"]:
        loan["note"] = sanitize_text(loan.get("note"))
    data["repayments"] = loaded.get("repayments", [])
    return data


def save_data(data: dict) -> None:
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def sort_entries(entries: list[dict]) -> list[dict]:
    return sorted(entries, key=lambda item: (item.get("date", ""), item.get("id", "")), reverse=True)


def new_id() -> str:
    return uuid4().hex[:10]


def remove_by_id(entries: list[dict], item_id: str) -> bool:
    before = len(entries)
    entries[:] = [item for item in entries if item.get("id") != item_id]
    return len(entries) < before


def update_local_item(
    entries: list[dict],
    item_id: str,
    label_key: str,
    label_value: str,
    amount_value: float,
    date_value: str,
) -> bool:
    for item in entries:
        if item.get("id") == item_id:
            item[label_key] = label_value
            item["amount"] = amount_value
            item["date"] = date_value
            return True
    return False


def sum_amount(entries: list[dict]) -> float:
    return round(sum(float(item.get("amount", 0.0) or 0.0) for item in entries), 2)


def unique_lenders(loans: list[dict], repayments: list[dict]) -> list[str]:
    names = set()
    for item in loans + repayments:
        lender = sanitize_text(item.get("lender", ""))
        if lender:
            names.add(lender)
    return sorted(names, key=str.lower)


def filter_repayments_by_lender(repayments: list[dict], lender: str | None) -> list[dict]:
    selected = sanitize_text(lender)
    if not selected:
        return repayments
    return [row for row in repayments if sanitize_text(row.get("lender", "")) == selected]


def build_chart_data(data: dict) -> dict:
    daily = defaultdict(
        lambda: {
            "acquisto_casa": 0.0,
            "ristrutturazione": 0.0,
            "prestiti": 0.0,
            "rimborsi": 0.0,
        }
    )

    for item in data["expenses"]["acquisto_casa"]:
        daily[item.get("date", "sconosciuto")[:10]]["acquisto_casa"] += float(item.get("amount", 0.0) or 0.0)

    for item in data["expenses"]["ristrutturazione"]:
        daily[item.get("date", "sconosciuto")[:10]]["ristrutturazione"] += float(item.get("amount", 0.0) or 0.0)

    for item in data["loans"]:
        daily[item.get("date", "sconosciuto")[:10]]["prestiti"] += float(item.get("amount", 0.0) or 0.0)

    for item in data["repayments"]:
        daily[item.get("date", "sconosciuto")[:10]]["rimborsi"] += float(item.get("amount", 0.0) or 0.0)

    labels = sorted(daily.keys())
    label_to_idx = {label: idx for idx, label in enumerate(labels)}

    def build_timeline_points(entries: list[dict], label_key: str) -> list[dict]:
        points = []
        for item in entries:
            amount = float(item.get("amount", 0.0) or 0.0)
            if amount <= 0:
                continue

            day_label = (item.get("date", "sconosciuto") or "sconosciuto")[:10]
            if day_label not in label_to_idx:
                continue

            voce = sanitize_text(item.get(label_key, ""))
            points.append(
                {
                    "x": label_to_idx[day_label],
                    "y": round(amount, 2),
                    "voce": voce,
                }
            )

        points.sort(key=lambda row: row["x"])
        return points

    totals = {
        "labels": ["Acquisto casa", "Ristrutturazione", "Prestiti ricevuti", "Rimborsi"],
        "values": [
            round(sum_amount(data["expenses"]["acquisto_casa"]), 2),
            round(sum_amount(data["expenses"]["ristrutturazione"]), 2),
            round(sum_amount(data["loans"]), 2),
            round(sum_amount(data["repayments"]), 2),
        ],
    }

    timeline = {
        "labels": labels,
        "acquistoCasa": build_timeline_points(data["expenses"]["acquisto_casa"], "description"),
        "ristrutturazione": build_timeline_points(data["expenses"]["ristrutturazione"], "description"),
        "prestiti": build_timeline_points(data["loans"], "lender"),
        "rimborsi": build_timeline_points(data["repayments"], "lender"),
    }

    return {
        "totals": totals,
        "timeline": timeline,
    }


def build_lender_balance(loans: list[dict], repayments: list[dict]) -> list[dict]:
    balances = defaultdict(float)

    for entry in loans:
        lender = entry.get("lender", "Sconosciuto")
        balances[lender] += float(entry.get("amount", 0.0) or 0.0)

    for entry in repayments:
        lender = entry.get("lender", "Sconosciuto")
        balances[lender] -= float(entry.get("amount", 0.0) or 0.0)

    result = []
    for lender, amount in sorted(balances.items(), key=lambda item: item[0].lower()):
        result.append({"lender": lender, "balance": round(amount, 2)})
    return result


def build_summary(data: dict) -> dict:
    acquisto_total = sum_amount(data["expenses"]["acquisto_casa"])
    ristr_total = sum_amount(data["expenses"]["ristrutturazione"])
    loans_total = sum_amount(data["loans"])
    repayments_total = sum_amount(data["repayments"])

    return {
        "acquisto_total": acquisto_total,
        "ristr_total": ristr_total,
        "spese_total": round(acquisto_total + ristr_total, 2),
        "loans_total": loans_total,
        "repayments_total": repayments_total,
        "debito_residuo": round(loans_total - repayments_total, 2),
        "cash_netto": round(loans_total - repayments_total - acquisto_total - ristr_total, 2),
        "lender_balance": build_lender_balance(data["loans"], data["repayments"]),
    }


def build_view_model(selected_repayment_lender: str | None = None) -> dict:
    data = load_data()
    data["expenses"]["acquisto_casa"] = sort_entries(data["expenses"]["acquisto_casa"])
    data["expenses"]["ristrutturazione"] = sort_entries(data["expenses"]["ristrutturazione"])
    data["loans"] = sort_entries(data["loans"])
    data["repayments"] = sort_entries(data["repayments"])

    selected_lender = sanitize_text(selected_repayment_lender)
    repayment_lenders = unique_lenders(data["loans"], data["repayments"])
    if selected_lender and selected_lender not in repayment_lenders:
        selected_lender = ""

    filtered_repayments = filter_repayments_by_lender(data["repayments"], selected_lender)

    return {
        "data": data,
        "summary": build_summary(data),
        "chart_data": build_chart_data(data),
        "today": date.today().isoformat(),
        "repayment_lenders": repayment_lenders,
        "selected_repayment_lender": selected_lender,
        "filtered_repayments": filtered_repayments,
    }


def list_managed_users() -> list[dict]:
    if not USE_DATABASE:
        return []

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT username, is_active, created_at, updated_at
                FROM users
                ORDER BY LOWER(username)
                """
            )
            rows = cur.fetchall()

    users: list[dict] = []
    for row in rows:
        users.append(
            {
                "username": sanitize_text(row.get("username")),
                "is_active": bool(row.get("is_active")),
                "created_at": row.get("created_at"),
                "updated_at": row.get("updated_at"),
            }
        )
    return users


def upsert_managed_user(username: str, password: str) -> str:
    if not USE_DATABASE:
        raise RuntimeError("Database non attivo")

    candidate_username = sanitize_text(username)
    if not candidate_username:
        raise ValueError("Username obbligatorio")
    if len(candidate_username) < 3:
        raise ValueError("Username troppo corto (min 3 caratteri)")
    if len(candidate_username) > 64:
        raise ValueError("Username troppo lungo (max 64 caratteri)")
    if not re.fullmatch(r"[A-Za-z0-9._@\-]+", candidate_username):
        raise ValueError("Username non valido (usa solo lettere, numeri e . _ @ -)")

    raw_password = password or ""
    if len(raw_password) < 8:
        raise ValueError("Password troppo corta (min 8 caratteri)")

    if normalize_username(candidate_username) == normalize_username(get_superadmin_username()):
        raise ValueError("Questo username e riservato all'utente amministratore da ambiente")

    password_hash = generate_password_hash(raw_password)

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO users (username, password_hash, is_active, updated_at)
                VALUES (%s, %s, TRUE, CURRENT_TIMESTAMP)
                ON CONFLICT (username)
                DO UPDATE SET
                    password_hash = EXCLUDED.password_hash,
                    is_active = TRUE,
                    updated_at = CURRENT_TIMESTAMP
                RETURNING username
                """,
                (candidate_username, password_hash),
            )
            row = cur.fetchone()
        conn.commit()

    return sanitize_text(row.get("username")) if isinstance(row, dict) else candidate_username


def delete_managed_user(username: str) -> bool:
    if not USE_DATABASE:
        return False

    candidate_username = sanitize_text(username)
    if not candidate_username:
        return False
    if normalize_username(candidate_username) == normalize_username(get_superadmin_username()):
        return False

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM users WHERE LOWER(username) = LOWER(%s)", (candidate_username,))
            deleted = cur.rowcount > 0
        if deleted:
            conn.commit()
    return deleted


@app.context_processor
def inject_auth_context() -> dict:
    return {
        "current_username": sanitize_text(session.get("username")),
        "can_manage_users": is_superadmin_session(),
    }


_db_bootstrap_done = False


@app.before_request
def bootstrap_database_if_needed():
    global _db_bootstrap_done
    if _db_bootstrap_done:
        return None
    ensure_database_ready()
    _db_bootstrap_done = True
    return None


@app.before_request
def require_login_for_app_routes():
    endpoint = request.endpoint or ""
    public_endpoints = {"login", "health_check", "static", "service_worker"}

    if endpoint in public_endpoints or endpoint.startswith("static"):
        return None
    if is_authenticated():
        return None

    next_path = request.full_path if request.query_string else request.path
    return redirect(url_for("login", next=next_path))


@app.route("/login", methods=["GET", "POST"])
def login():
    if is_authenticated():
        return redirect(url_for("index"))

    error = None
    next_path = sanitize_text(request.args.get("next"))

    if request.method == "POST":
        username = sanitize_text(request.form.get("username"))
        password = request.form.get("password") or ""
        form_next = sanitize_text(request.form.get("next"))

        auth_result = authenticate_login(username, password)
        if auth_result:
            session.clear()
            # Keep the auth session across browser restarts until logout or cookie expiry.
            session.permanent = True
            session["authenticated"] = True
            session["username"] = sanitize_text(auth_result.get("username"))
            session["role"] = sanitize_text(auth_result.get("role"))
            target = form_next if _is_safe_next_path(form_next) else next_path
            return redirect(target if _is_safe_next_path(target) else url_for("index"))

        error = "Credenziali non valide."

    return render_template("login.html", error=error, next_path=next_path)


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/admin/users", methods=["GET"])
def admin_users_page():
    if not is_superadmin_session():
        return Response("Forbidden", status=403)

    users: list[dict] = []
    if USE_DATABASE:
        try:
            users = list_managed_users()
        except Exception as exc:
            session["admin_users_error"] = f"Errore caricamento utenti: {exc}"

    return render_template(
        "admin_users.html",
        users=users,
        user_management_available=USE_DATABASE,
        notice=sanitize_text(session.pop("admin_users_notice", "")),
        error=sanitize_text(session.pop("admin_users_error", "")),
    )


@app.route("/admin/users", methods=["POST"])
def admin_upsert_user():
    if not is_superadmin_session():
        return Response("Forbidden", status=403)

    if not USE_DATABASE:
        session["admin_users_error"] = "Gestione utenti disponibile solo con database attivo."
        return redirect(url_for("admin_users_page"))

    username = sanitize_text(request.form.get("username"))
    password = request.form.get("password") or ""

    try:
        saved_username = upsert_managed_user(username, password)
        session["admin_users_notice"] = f"Utente '{saved_username}' salvato correttamente."
    except Exception as exc:
        session["admin_users_error"] = str(exc)

    return redirect(url_for("admin_users_page"))


@app.route("/admin/users/delete", methods=["POST"])
def admin_delete_user():
    if not is_superadmin_session():
        return Response("Forbidden", status=403)

    username = sanitize_text(request.form.get("username"))
    if not username:
        session["admin_users_error"] = "Username mancante."
        return redirect(url_for("admin_users_page"))

    if normalize_username(username) == normalize_username(get_superadmin_username()):
        session["admin_users_error"] = "L'utente amministratore da ambiente non puo essere rimosso."
        return redirect(url_for("admin_users_page"))

    try:
        deleted = delete_managed_user(username)
        if deleted:
            session["admin_users_notice"] = f"Utente '{username}' eliminato."
        else:
            session["admin_users_error"] = "Utente non trovato."
    except Exception as exc:
        session["admin_users_error"] = str(exc)

    return redirect(url_for("admin_users_page"))


@app.route("/", methods=["GET"])
def index():
    vm = build_view_model(request.args.get("repayment_lender"))
    return render_template(
        "index.html",
        data=vm["data"],
        summary=vm["summary"],
        chart_data=vm["chart_data"],
        today=vm["today"],
        repayment_lenders=vm["repayment_lenders"],
        selected_repayment_lender=vm["selected_repayment_lender"],
        filtered_repayments=vm["filtered_repayments"],
        error=None,
    )


def render_index_error(error_message: str):
    vm = build_view_model()
    return render_template(
        "index.html",
        data=vm["data"],
        summary=vm["summary"],
        chart_data=vm["chart_data"],
        today=vm["today"],
        repayment_lenders=vm["repayment_lenders"],
        selected_repayment_lender=vm["selected_repayment_lender"],
        filtered_repayments=vm["filtered_repayments"],
        error=error_message,
    )


@app.route("/health", methods=["GET"])
def health_check():
    return Response("OK", status=200, mimetype="text/plain")


# ---------------------------------------------------------------------------
# AI Agent – Gemini native Function Calling
# ---------------------------------------------------------------------------

AGENT_SYSTEM_INSTRUCTION = (
    "Sei l'assistente AI dell'app Home13, dedicata alla gestione di spese casa, "
    "prestiti ricevuti e rimborsi.\n\n"
    "PRIORITÀ ASSOLUTE (in ordine):\n"
    "1) Accuratezza operativa: usa sempre le funzioni disponibili per leggere/modificare dati.\n"
    "2) Sicurezza utente: non eseguire mai cancellazioni/modifiche senza conferma esplicita.\n"
    "3) Linguaggio utente: niente dettagli tecnici interni, risposte chiare e brevi.\n\n"
    "DOMINIO DATI:\n"
    "- Spese acquisto casa (category=acquisto_casa): notaio, rogito, agenzia, imposte, caparra, mutuo, compravendita.\n"
    "- Spese ristrutturazione (category=ristrutturazione): lavori, materiali, mobili, elettrodomestici, impianti, artigiani.\n"
    "- Prestiti ricevuti (section=loans): denaro ricevuto da persone.\n"
    "- Rimborsi (section=repayments): denaro restituito ai prestatori.\n\n"
    "REGOLE DI RISPOSTA:\n"
    "- Rispondi sempre in italiano, tono naturale, frasi concise e orientate all'azione.\n"
    "- Quando mostri date all'utente, usa sempre il formato italiano GG/MM/AAAA.\n"
    "- NON menzionare mai ID, codici, chiavi interne, nomi di campo o dettagli tecnici.\n"
    "- NON mostrare mai etichette tecniche come acquisto_casa/loans/repayments: usa sempre nomi umani con iniziale maiuscola.\n"
    "- Per identificare una voce usa solo riferimenti umani: data, importo, descrizione, prestatore.\n"
    "- Non inventare mai dati: se non trovi una voce, dichiaralo chiaramente.\n"
    "- Se mancano dati obbligatori, chiedi solo il minimo indispensabile.\n\n"
    "COERENZA NUMERICA (OBBLIGATORIA):\n"
    "- Per richieste su un prestatore specifico ('saldo', 'totale con X', 'quanto devo a X'), usa SEMPRE get_lender_metrics.\n"
    "- Per domande di verifica/conferma numerica ('e giusto?', 'confermi?', 'corretto?') usa SEMPRE verify_lender_amount prima di rispondere.\n"
    "- NON riutilizzare importi dalla memoria conversazionale senza ricalcolo via funzione.\n\n"
    "GESTIONE DATE (obbligatoria):\n"
    "- Nei parametri funzione usa sempre formato ISO YYYY-MM-DD.\n"
    "- Converte automaticamente: 'oggi'={today}, 'ieri'={yesterday}, 'domani'={tomorrow}, "
    "'dopodomani'={day_after_tomorrow}, 'l'altro ieri'={day_before_yesterday}.\n"
    "- Se l'utente ha già indicato una data (esplicita o relativa), NON richiederla di nuovo.\n"
    "- Se la data non è specificata, puoi assumere oggi solo per nuove registrazioni; "
    "per modifiche/cancellazioni chiedi conferma dei riferimenti.\n\n"
    "DEDUZIONE CATEGORIA:\n"
    "- Se il contesto è acquisto immobile/atto/rogito/mutuo -> acquisto_casa.\n"
    "- Altrimenti, per spese di lavori/forniture/casa -> ristrutturazione.\n\n"
    "PROTOCOLLO OPERATIVO:\n"
    "- AGGIUNTA: se i dati minimi sono presenti, esegui direttamente la funzione di inserimento.\n"
    "- MODIFICA: prima cerca con search_entries, mostra la voce candidata in forma umana, "
    "chiedi conferma dei nuovi valori, poi aggiorna.\n"
    "- CANCELLAZIONE SINGOLA: prima cerca con search_entries, mostra cosa verrà eliminato, "
    "chiedi conferma esplicita, poi elimina.\n"
    "- CANCELLAZIONE MASSIVA: usa search_entries per quantificare le voci coinvolte, "
    "chiedi conferma esplicita, poi usa delete_all_entries con confirm=true.\n"
    "- RICHIESTE DI TOTALI/INTERVALLI: usa search_entries con date_from/date_to e include_entries=false, "
    "poi rispondi dal campo total_amount.\n\n"
    "GESTIONE AMBIGUITÀ:\n"
    "- Se trovi più voci simili, non scegliere in autonomia: chiedi un disambiguatore (data/importo/descrizione/prestatore).\n"
    "- Fai una sola domanda di chiarimento per volta, la più utile a sbloccare l'azione.\n"
    "- Se l'utente conferma con 'si/sì', procedi senza ripetere domande già risolte.\n\n"
    "INFERENZA DA HISTORY RECENTE (IMPORTANTE):\n"
    "- La history dei turni precedenti è SEMPRE disponibile. Usala per inferire il contesto implicito.\n"
    "- Strategia: Se l'utente usa pronomi dimostrativi ('quell'importo', 'quella spesa', 'l'ultima voce', 'quel prestito'), "
    "estrai il contesto dal turno precedente dove è stata registrata/cercata una voce.\n"
    "- REGOLA CHIAVE: Nel turno precedente (model response), estrai ID e attributi della voce menzionata "
    "(es: se ho detto 'Ho registrato l'acquisto della TV da 1.200 euro (ID: abc123)', e l'utente dice 'modifica quell'importo a 1.300', "
    "usa ID=abc123 direttamente senza chiedere quale spesa).\n"
    "- Se nel turno modello più recente c'è stata UNA ricerca che ha trovato UNA SOLA voce, considera quella come il target implicito "
    "per qualsiasi azione di modifica/cancellazione nel turno attuale.\n"
    "- NON sprecare turni con disambiguatori quando il contesto dalla history lo rende OVVIO.\n\n"
    "POST-OPERAZIONE:\n"
    "- Dopo aggiunta/modifica/cancellazione riuscita, termina il messaggio con [REFRESH] esattamente una volta.\n"
    "- Non usare [REFRESH] se nessuna modifica dati è stata effettuata.\n"
)

_ISO_DATE_PATTERN = r"^\d{4}-\d{2}-\d{2}$"

AGENT_TOOL_DECLARATIONS: dict = {
    "function_declarations": [
        {
            "name": "get_summary",
            "description": (
                "Restituisce il riepilogo finanziario aggregato (totali, debito residuo, saldo per prestatore)."
            ),
            "parameters": {
                "type": "object",
                "properties": {},
                "required": [],
            },
        },
        {
            "name": "search_entries",
            "description": (
                "Ricerca voci esistenti. Usalo per disambiguare prima di update/delete "
                "e per rispondere a richieste su elenco/somme/intervalli."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "section": {
                        "type": "string",
                        "enum": ["acquisto_casa", "ristrutturazione", "loans", "repayments", "all"],
                        "description": "Sezione in cui cercare.",
                    },
                    "description": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Filtro testuale descrizione (solo spese).",
                    },
                    "lender": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Filtro per prestatore (prestiti/rimborsi).",
                    },
                    "amount": {
                        "type": "number",
                        "minimum": 0.01,
                        "description": "Importo esatto (euro).",
                    },
                    "date": {
                        "type": "string",
                        "pattern": _ISO_DATE_PATTERN,
                        "description": "Data singola YYYY-MM-DD (non usarla insieme a date_from/date_to).",
                    },
                    "date_from": {
                        "type": "string",
                        "pattern": _ISO_DATE_PATTERN,
                        "description": "Data inizio intervallo inclusa YYYY-MM-DD.",
                    },
                    "date_to": {
                        "type": "string",
                        "pattern": _ISO_DATE_PATTERN,
                        "description": "Data fine intervallo inclusa YYYY-MM-DD.",
                    },
                    "include_entries": {
                        "type": "boolean",
                        "description": "Se false restituisce solo count/total_amount senza elenco dettagliato.",
                    },
                },
                "required": ["section"],
            },
        },
        {
            "name": "get_lender_metrics",
            "description": (
                "Restituisce metriche affidabili per un singolo prestatore: totale prestiti ricevuti, "
                "totale rimborsi e saldo residuo (prestiti-rimborsi)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "lender": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Nome del prestatore da analizzare.",
                    },
                },
                "required": ["lender"],
            },
        },
        {
            "name": "verify_lender_amount",
            "description": (
                "Verifica se un importo citato dall'utente e coerente con i dati reali "
                "di un prestatore (saldo o totali)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "lender": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Nome del prestatore da verificare.",
                    },
                    "expected_amount": {
                        "type": "number",
                        "minimum": 0,
                        "description": "Importo dichiarato dall'utente da verificare.",
                    },
                    "metric": {
                        "type": "string",
                        "enum": ["balance", "loans_total", "repayments_total"],
                        "description": "Tipo di metrica da confrontare.",
                    },
                    "tolerance": {
                        "type": "number",
                        "minimum": 0,
                        "maximum": 1000,
                        "description": "Tolleranza ammessa in euro (default 0.01).",
                    },
                },
                "required": ["lender", "expected_amount", "metric"],
            },
        },
        {
            "name": "add_expense",
            "description": "Registra una nuova spesa in acquisto_casa o ristrutturazione.",
            "parameters": {
                "type": "object",
                "properties": {
                    "category": {
                        "type": "string",
                        "enum": ["acquisto_casa", "ristrutturazione"],
                        "description": "Categoria della spesa.",
                    },
                    "description": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Descrizione utente della spesa.",
                    },
                    "date": {
                        "type": "string",
                        "pattern": _ISO_DATE_PATTERN,
                        "description": "Data operazione in formato YYYY-MM-DD.",
                    },
                    "amount": {
                        "type": "number",
                        "minimum": 0.01,
                        "description": "Importo positivo in euro.",
                    },
                },
                "required": ["category", "description", "date", "amount"],
            },
        },
        {
            "name": "add_loan",
            "description": "Registra un nuovo prestito ricevuto.",
            "parameters": {
                "type": "object",
                "properties": {
                    "lender": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Nome prestatore.",
                    },
                    "date": {
                        "type": "string",
                        "pattern": _ISO_DATE_PATTERN,
                        "description": "Data operazione in formato YYYY-MM-DD.",
                    },
                    "amount": {
                        "type": "number",
                        "minimum": 0.01,
                        "description": "Importo positivo in euro.",
                    },
                    "note": {
                        "type": "string",
                        "description": "Nota opzionale.",
                    },
                },
                "required": ["lender", "date", "amount"],
            },
        },
        {
            "name": "add_repayment",
            "description": "Registra un nuovo rimborso effettuato a un prestatore.",
            "parameters": {
                "type": "object",
                "properties": {
                    "lender": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Nome prestatore rimborsato.",
                    },
                    "date": {
                        "type": "string",
                        "pattern": _ISO_DATE_PATTERN,
                        "description": "Data operazione in formato YYYY-MM-DD.",
                    },
                    "amount": {
                        "type": "number",
                        "minimum": 0.01,
                        "description": "Importo positivo in euro.",
                    },
                },
                "required": ["lender", "date", "amount"],
            },
        },
        {
            "name": "delete_expense",
            "description": (
                "Elimina una spesa esistente. Da usare solo dopo search_entries e conferma esplicita utente."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "id": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Identificatore interno voce (mai da mostrare all'utente).",
                    },
                    "category": {
                        "type": "string",
                        "enum": ["acquisto_casa", "ristrutturazione"],
                        "description": "Categoria della voce da eliminare.",
                    },
                },
                "required": ["id", "category"],
            },
        },
        {
            "name": "delete_loan",
            "description": "Elimina un prestito esistente dopo ricerca e conferma esplicita utente.",
            "parameters": {
                "type": "object",
                "properties": {
                    "id": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Identificatore interno voce (mai da mostrare all'utente).",
                    },
                },
                "required": ["id"],
            },
        },
        {
            "name": "delete_repayment",
            "description": "Elimina un rimborso esistente dopo ricerca e conferma esplicita utente.",
            "parameters": {
                "type": "object",
                "properties": {
                    "id": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Identificatore interno voce (mai da mostrare all'utente).",
                    },
                },
                "required": ["id"],
            },
        },
        {
            "name": "delete_all_entries",
            "description": (
                "Svuota completamente una sezione. Da usare solo dopo conferma esplicita dell'utente."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "section": {
                        "type": "string",
                        "enum": ["acquisto_casa", "ristrutturazione", "loans", "repayments"],
                        "description": "Sezione da eliminare integralmente.",
                    },
                    "confirm": {
                        "type": "boolean",
                        "description": "Deve essere true solo dopo consenso esplicito utente.",
                    },
                },
                "required": ["section", "confirm"],
            },
        },
        {
            "name": "update_expense",
            "description": (
                "Modifica una spesa esistente. Obbligatorio fare search_entries prima e poi conferma utente."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "id": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Identificatore interno voce (mai da mostrare all'utente).",
                    },
                    "category": {
                        "type": "string",
                        "enum": ["acquisto_casa", "ristrutturazione"],
                        "description": "Categoria della spesa.",
                    },
                    "description": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Nuova descrizione (opzionale).",
                    },
                    "date": {
                        "type": "string",
                        "pattern": _ISO_DATE_PATTERN,
                        "description": "Nuova data YYYY-MM-DD (opzionale).",
                    },
                    "amount": {
                        "type": "number",
                        "minimum": 0.01,
                        "description": "Nuovo importo positivo (opzionale).",
                    },
                },
                "required": ["id", "category"],
            },
        },
        {
            "name": "update_loan",
            "description": (
                "Modifica un prestito esistente. Obbligatorio fare search_entries prima e poi conferma utente."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "id": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Identificatore interno voce (mai da mostrare all'utente).",
                    },
                    "lender": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Nuovo prestatore (opzionale).",
                    },
                    "date": {
                        "type": "string",
                        "pattern": _ISO_DATE_PATTERN,
                        "description": "Nuova data YYYY-MM-DD (opzionale).",
                    },
                    "amount": {
                        "type": "number",
                        "minimum": 0.01,
                        "description": "Nuovo importo positivo (opzionale).",
                    },
                    "note": {
                        "type": "string",
                        "description": "Nuova nota (opzionale).",
                    },
                },
                "required": ["id"],
            },
        },
        {
            "name": "update_repayment",
            "description": (
                "Modifica un rimborso esistente. Obbligatorio fare search_entries prima e poi conferma utente."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "id": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Identificatore interno voce (mai da mostrare all'utente).",
                    },
                    "lender": {
                        "type": "string",
                        "minLength": 1,
                        "description": "Nuovo prestatore (opzionale).",
                    },
                    "date": {
                        "type": "string",
                        "pattern": _ISO_DATE_PATTERN,
                        "description": "Nuova data YYYY-MM-DD (opzionale).",
                    },
                    "amount": {
                        "type": "number",
                        "minimum": 0.01,
                        "description": "Nuovo importo positivo (opzionale).",
                    },
                },
                "required": ["id"],
            },
        },
    ]
}


def _agent_search_section(
    entries: list[dict],
    section_name: str,
    text_key: str,
    text_filter: str | None,
    amount_filter: float | None,
    date_filter: str | None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> list[dict]:
    results = []
    for entry in entries:
        if text_filter and normalize_text(text_filter) not in normalize_text(entry.get(text_key, "")):
            continue
        if amount_filter is not None and abs(float(entry.get("amount", 0.0)) - amount_filter) > 0.01:
            continue
        entry_date = entry.get("date", "")
        if date_filter and entry_date != date_filter:
            continue
        if date_from and entry_date < date_from:
            continue
        if date_to and entry_date > date_to:
            continue
        results.append({**entry, "section": section_name})
    return results


def agent_fn_get_summary() -> dict:
    summary = build_summary(load_data())
    return {
        "success": True,
        "summary": {
            "spese_acquisto_casa": f"{format_euro(summary['acquisto_total'])} €",
            "spese_ristrutturazione": f"{format_euro(summary['ristr_total'])} €",
            "spese_totali": f"{format_euro(summary['spese_total'])} €",
            "prestiti_ricevuti": f"{format_euro(summary['loans_total'])} €",
            "rimborsi_effettuati": f"{format_euro(summary['repayments_total'])} €",
            "debito_residuo": f"{format_euro(summary['debito_residuo'])} €",
            "saldo_per_prestatore": summary["lender_balance"],
        },
    }


def _compute_lender_metrics(lender_name: str) -> dict:
    candidate = sanitize_text(lender_name)
    if not candidate:
        return {
            "found": False,
            "lender": "",
            "loans_total": 0.0,
            "repayments_total": 0.0,
            "balance": 0.0,
        }

    normalized_candidate = normalize_text(candidate)
    data = load_data()

    loans_total = 0.0
    repayments_total = 0.0
    matched_names: list[str] = []

    for item in data["loans"]:
        lender = sanitize_text(item.get("lender", ""))
        if lender and normalize_text(lender) == normalized_candidate:
            loans_total += float(item.get("amount", 0.0) or 0.0)
            matched_names.append(lender)

    for item in data["repayments"]:
        lender = sanitize_text(item.get("lender", ""))
        if lender and normalize_text(lender) == normalized_candidate:
            repayments_total += float(item.get("amount", 0.0) or 0.0)
            matched_names.append(lender)

    canonical_lender = matched_names[0] if matched_names else candidate
    loans_total = round(loans_total, 2)
    repayments_total = round(repayments_total, 2)
    balance = round(loans_total - repayments_total, 2)

    return {
        "found": bool(matched_names),
        "lender": canonical_lender,
        "loans_total": loans_total,
        "repayments_total": repayments_total,
        "balance": balance,
    }


def agent_fn_get_lender_metrics(args: dict) -> dict:
    lender = sanitize_text(args.get("lender", ""))
    if not lender:
        return {"success": False, "error": "Prestatore mancante."}

    metrics = _compute_lender_metrics(lender)
    return {
        "success": True,
        "found": metrics["found"],
        "lender": metrics["lender"],
        "metrics": {
            "loans_total": metrics["loans_total"],
            "repayments_total": metrics["repayments_total"],
            "balance": metrics["balance"],
            "loans_total_formatted": f"{format_euro(metrics['loans_total'])} €",
            "repayments_total_formatted": f"{format_euro(metrics['repayments_total'])} €",
            "balance_formatted": f"{format_euro(metrics['balance'])} €",
        },
    }


def agent_fn_verify_lender_amount(args: dict) -> dict:
    lender = sanitize_text(args.get("lender", ""))
    metric = sanitize_text(args.get("metric", "")).casefold()
    if not lender:
        return {"success": False, "error": "Prestatore mancante."}
    if metric not in {"balance", "loans_total", "repayments_total"}:
        return {"success": False, "error": "Metrica non valida."}

    expected_raw = args.get("expected_amount")
    try:
        expected_amount = round(float(expected_raw), 2)
    except (TypeError, ValueError):
        return {"success": False, "error": "Importo atteso non valido."}

    tolerance_raw = args.get("tolerance")
    try:
        tolerance = float(tolerance_raw) if tolerance_raw is not None else 0.01
    except (TypeError, ValueError):
        tolerance = 0.01
    tolerance = max(0.0, min(1000.0, tolerance))

    metrics = _compute_lender_metrics(lender)
    if not metrics["found"]:
        return {
            "success": True,
            "found": False,
            "lender": lender,
            "metric": metric,
            "is_match": False,
            "message": "Prestatore non trovato.",
        }

    actual_amount = round(float(metrics[metric]), 2)
    difference = round(actual_amount - expected_amount, 2)
    is_match = abs(difference) <= tolerance

    return {
        "success": True,
        "found": True,
        "lender": metrics["lender"],
        "metric": metric,
        "expected_amount": expected_amount,
        "actual_amount": actual_amount,
        "difference": difference,
        "tolerance": round(tolerance, 2),
        "is_match": is_match,
        "expected_amount_formatted": f"{format_euro(expected_amount)} €",
        "actual_amount_formatted": f"{format_euro(actual_amount)} €",
        "difference_formatted": f"{format_euro(abs(difference))} €",
    }


def agent_fn_search_entries(args: dict) -> dict:
    section = canonical_section(sanitize_text(args.get("section", "all")), allow_all=True)
    if not section:
        return {"success": False, "error": "Sezione non valida."}

    desc_filter = sanitize_text(args.get("description")) or None
    lender_filter = sanitize_text(args.get("lender")) or None
    amount_filter = float(args["amount"]) if args.get("amount") is not None else None
    date_filter = sanitize_text(args.get("date")) or None
    date_from = sanitize_text(args.get("date_from")) or None
    date_to = sanitize_text(args.get("date_to")) or None
    include_entries_raw = args.get("include_entries", True)
    if isinstance(include_entries_raw, str):
        include_entries = sanitize_text(include_entries_raw).casefold() not in {"false", "0", "no", "off"}
    else:
        include_entries = bool(include_entries_raw)

    data = load_data()
    results: list[dict] = []
    target_sections = (
        ["acquisto_casa", "ristrutturazione", "loans", "repayments"]
        if section == "all"
        else [section]
    )

    for sec in target_sections:
        if sec in ("acquisto_casa", "ristrutturazione"):
            results += _agent_search_section(
                data["expenses"].get(sec, []), sec, "description",
                desc_filter, amount_filter, date_filter, date_from, date_to,
            )
        elif sec == "loans":
            results += _agent_search_section(
                data["loans"], "loans", "lender",
                lender_filter, amount_filter, date_filter, date_from, date_to,
            )
        elif sec == "repayments":
            results += _agent_search_section(
                data["repayments"], "repayments", "lender",
                lender_filter, amount_filter, date_filter, date_from, date_to,
            )

    total = round(sum(float(e.get("amount", 0.0)) for e in results), 2)
    response = {"success": True, "count": len(results), "total_amount": total}
    if include_entries:
        response["entries"] = results
    return response


def _normalize_client_history(raw_history: list[dict], max_turns: int) -> list[dict]:
    normalized: list[dict] = []
    for turn in raw_history:
        if not isinstance(turn, dict):
            continue
        role = turn.get("role")
        if role not in {"user", "model"}:
            continue
        parts = turn.get("parts")
        if not isinstance(parts, list):
            continue

        text_parts = []
        for part in parts:
            if not isinstance(part, dict):
                continue
            text = sanitize_text((part or {}).get("text", ""))
            if text:
                text_parts.append({"text": text})

        if text_parts:
            normalized.append({"role": role, "parts": text_parts})

    if len(normalized) > max_turns:
        normalized = normalized[-max_turns:]

    if normalized and normalized[0].get("role") == "model":
        normalized = normalized[1:]

    return normalized


def _build_client_safe_history(history: list[dict], max_turns: int) -> list[dict]:
    return _normalize_client_history(history, max_turns)


def _is_same_last_user_turn(history: list[dict], message: str) -> bool:
    if not history:
        return False
    last_turn = history[-1]
    if not isinstance(last_turn, dict) or last_turn.get("role") != "user":
        return False

    last_text = "\n".join(
        sanitize_text((part or {}).get("text", ""))
        for part in (last_turn.get("parts") or [])
        if isinstance(part, dict) and (part or {}).get("text") is not None
    ).strip()

    return last_text == sanitize_text(message)


def agent_fn_add_expense(args: dict) -> dict:
    category = canonical_section(sanitize_text(args.get("category", "")), allow_all=False)
    if category not in {"acquisto_casa", "ristrutturazione"}:
        return {"success": False, "error": "Categoria non valida."}
    description = capitalize_first(sanitize_text(args.get("description", ""))) or "Spesa"
    try:
        item_date = parse_iso_date(sanitize_text(args.get("date", ""))).isoformat()
    except ValueError:
        return {"success": False, "error": "Data non valida."}
    amount_raw = args.get("amount")
    try:
        amount = float(amount_raw)
        if amount <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return {"success": False, "error": "Importo non valido."}

    item = {"id": new_id(), "date": item_date, "description": description, "amount": amount}
    if USE_DATABASE:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO expenses (id, category, operation_date, description, amount) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (item["id"], category, item["date"], item["description"], item["amount"]),
                )
            conn.commit()
    else:
        data = load_data()
        data["expenses"][category].append(item)
        save_data(data)
    return {"success": True, "refresh": True, "item": item, "category": category}


def agent_fn_add_loan(args: dict) -> dict:
    lender = capitalize_first(sanitize_text(args.get("lender", "")))
    if not lender:
        return {"success": False, "error": "Nome prestatore mancante."}
    try:
        item_date = parse_iso_date(sanitize_text(args.get("date", ""))).isoformat()
    except ValueError:
        return {"success": False, "error": "Data non valida."}
    try:
        amount = float(args.get("amount", 0))
        if amount <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return {"success": False, "error": "Importo non valido."}

    item = {
        "id": new_id(),
        "date": item_date,
        "lender": lender,
        "note": sanitize_text(args.get("note", "")),
        "amount": amount,
    }
    if USE_DATABASE:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO loans (id, operation_date, lender, note, amount) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (item["id"], item["date"], item["lender"], item["note"], item["amount"]),
                )
            conn.commit()
    else:
        data = load_data()
        data["loans"].append(item)
        save_data(data)
    return {"success": True, "refresh": True, "item": item}


def agent_fn_add_repayment(args: dict) -> dict:
    lender = capitalize_first(sanitize_text(args.get("lender", "")))
    if not lender:
        return {"success": False, "error": "Nome prestatore mancante."}
    try:
        item_date = parse_iso_date(sanitize_text(args.get("date", ""))).isoformat()
    except ValueError:
        return {"success": False, "error": "Data non valida."}
    try:
        amount = float(args.get("amount", 0))
        if amount <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return {"success": False, "error": "Importo non valido."}

    item = {"id": new_id(), "date": item_date, "lender": lender, "amount": amount}
    if USE_DATABASE:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO repayments (id, operation_date, lender, amount) VALUES (%s, %s, %s, %s)",
                    (item["id"], item["date"], item["lender"], item["amount"]),
                )
            conn.commit()
    else:
        data = load_data()
        data["repayments"].append(item)
        save_data(data)
    return {"success": True, "refresh": True, "item": item}


def agent_fn_delete_expense(args: dict) -> dict:
    item_id = sanitize_text(args.get("id", ""))
    category = canonical_section(sanitize_text(args.get("category", "")), allow_all=False)
    if not item_id or category not in {"acquisto_casa", "ristrutturazione"}:
        return {"success": False, "error": "Voce o categoria mancante/non valida."}
    deleted = delete_ai_operation(
        "delete_expense_acquisto_casa" if category == "acquisto_casa" else "delete_expense_ristrutturazione",
        item_id,
    )
    if deleted:
        return {"success": True, "refresh": True}
    return {"success": False, "error": "Voce non trovata o già eliminata."}


def agent_fn_delete_loan(args: dict) -> dict:
    item_id = sanitize_text(args.get("id", ""))
    if not item_id:
        return {"success": False, "error": "Voce mancante."}
    deleted = delete_ai_operation("delete_loan", item_id)
    if deleted:
        return {"success": True, "refresh": True}
    return {"success": False, "error": "Prestito non trovato o già eliminato."}


def agent_fn_delete_repayment(args: dict) -> dict:
    item_id = sanitize_text(args.get("id", ""))
    if not item_id:
        return {"success": False, "error": "Voce mancante."}
    deleted = delete_ai_operation("delete_repayment", item_id)
    if deleted:
        return {"success": True, "refresh": True}
    return {"success": False, "error": "Rimborso non trovato o già eliminato."}


def _is_explicit_confirmation(value) -> bool:
    if value is True:
        return True
    text = sanitize_text(str(value)).casefold()
    return text in {"true", "1", "yes", "ok", "conferma", "si", "sì"}


def agent_fn_delete_all_entries(args: dict) -> dict:
    section = canonical_section(sanitize_text(args.get("section", "")), allow_all=False)
    allowed_sections = {"acquisto_casa", "ristrutturazione", "loans", "repayments"}
    if section not in allowed_sections:
        return {"success": False, "error": "Sezione non valida."}

    if not _is_explicit_confirmation(args.get("confirm")):
        return {
            "success": False,
            "error": "Conferma esplicita obbligatoria. Ripeti con confirm=true dopo consenso utente.",
        }

    deleted_count = 0

    if USE_DATABASE:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                if section in {"acquisto_casa", "ristrutturazione"}:
                    cur.execute("DELETE FROM expenses WHERE category = %s", (section,))
                elif section == "loans":
                    cur.execute("DELETE FROM loans")
                else:
                    cur.execute("DELETE FROM repayments")
                deleted_count = cur.rowcount
            if deleted_count > 0:
                conn.commit()
    else:
        data = load_data()
        if section in {"acquisto_casa", "ristrutturazione"}:
            deleted_count = len(data["expenses"][section])
            data["expenses"][section] = []
        elif section == "loans":
            deleted_count = len(data["loans"])
            data["loans"] = []
        else:
            deleted_count = len(data["repayments"])
            data["repayments"] = []

        if deleted_count > 0:
            save_data(data)

    return {
        "success": True,
        "refresh": True,
        "section": section,
        "deleted_count": int(deleted_count),
    }


def _apply_update_fields(item: dict, args: dict, text_key: str) -> None:
    if args.get(text_key):
        item[text_key] = capitalize_first(sanitize_text(args[text_key]))
    if args.get("date"):
        try:
            item["date"] = parse_iso_date(sanitize_text(args["date"])).isoformat()
        except ValueError:
            pass
    if args.get("amount") is not None:
        try:
            val = float(args["amount"])
            if val > 0:
                item["amount"] = val
        except (TypeError, ValueError):
            pass


def agent_fn_update_expense(args: dict) -> dict:
    item_id = sanitize_text(args.get("id", ""))
    category = canonical_section(sanitize_text(args.get("category", "")), allow_all=False)
    if not item_id or category not in {"acquisto_casa", "ristrutturazione"}:
        return {"success": False, "error": "Voce o categoria mancante/non valida."}

    if USE_DATABASE:
        sets = []
        params: list = []
        if args.get("description"):
            sets.append("description = %s")
            params.append(capitalize_first(sanitize_text(args["description"])))
        if args.get("date"):
            try:
                sets.append("operation_date = %s")
                params.append(parse_iso_date(sanitize_text(args["date"])).isoformat())
            except ValueError:
                pass
        if args.get("amount") is not None:
            try:
                val = float(args["amount"])
                if val > 0:
                    sets.append("amount = %s")
                    params.append(val)
            except (TypeError, ValueError):
                pass
        if not sets:
            return {"success": False, "error": "Nessun campo da aggiornare."}
        params += [item_id, category]
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"UPDATE expenses SET {', '.join(sets)} WHERE id = %s AND category = %s",
                    params,
                )
                updated = cur.rowcount > 0
            if updated:
                conn.commit()
        if updated:
            return {"success": True, "refresh": True}
        return {"success": False, "error": "Voce non trovata."}

    data = load_data()
    for item in data["expenses"][category]:
        if item.get("id") == item_id:
            _apply_update_fields(item, args, "description")
            save_data(data)
            return {"success": True, "refresh": True}
    return {"success": False, "error": "Voce non trovata."}


def agent_fn_update_loan(args: dict) -> dict:
    item_id = sanitize_text(args.get("id", ""))
    if not item_id:
        return {"success": False, "error": "Voce mancante."}

    if USE_DATABASE:
        sets = []
        params: list = []
        if args.get("lender"):
            sets.append("lender = %s")
            params.append(capitalize_first(sanitize_text(args["lender"])))
        if args.get("date"):
            try:
                sets.append("operation_date = %s")
                params.append(parse_iso_date(sanitize_text(args["date"])).isoformat())
            except ValueError:
                pass
        if args.get("amount") is not None:
            try:
                val = float(args["amount"])
                if val > 0:
                    sets.append("amount = %s")
                    params.append(val)
            except (TypeError, ValueError):
                pass
        if args.get("note") is not None:
            sets.append("note = %s")
            params.append(sanitize_text(args["note"]))
        if not sets:
            return {"success": False, "error": "Nessun campo da aggiornare."}
        params.append(item_id)
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"UPDATE loans SET {', '.join(sets)} WHERE id = %s", params)
                updated = cur.rowcount > 0
            if updated:
                conn.commit()
        if updated:
            return {"success": True, "refresh": True}
        return {"success": False, "error": "Prestito non trovato."}

    data = load_data()
    for item in data["loans"]:
        if item.get("id") == item_id:
            _apply_update_fields(item, args, "lender")
            if args.get("note") is not None:
                item["note"] = sanitize_text(args["note"])
            save_data(data)
            return {"success": True, "refresh": True}
    return {"success": False, "error": "Prestito non trovato."}


def agent_fn_update_repayment(args: dict) -> dict:
    item_id = sanitize_text(args.get("id", ""))
    if not item_id:
        return {"success": False, "error": "Voce mancante."}

    if USE_DATABASE:
        sets = []
        params: list = []
        if args.get("lender"):
            sets.append("lender = %s")
            params.append(capitalize_first(sanitize_text(args["lender"])))
        if args.get("date"):
            try:
                sets.append("operation_date = %s")
                params.append(parse_iso_date(sanitize_text(args["date"])).isoformat())
            except ValueError:
                pass
        if args.get("amount") is not None:
            try:
                val = float(args["amount"])
                if val > 0:
                    sets.append("amount = %s")
                    params.append(val)
            except (TypeError, ValueError):
                pass
        if not sets:
            return {"success": False, "error": "Nessun campo da aggiornare."}
        params.append(item_id)
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"UPDATE repayments SET {', '.join(sets)} WHERE id = %s", params)
                updated = cur.rowcount > 0
            if updated:
                conn.commit()
        if updated:
            return {"success": True, "refresh": True}
        return {"success": False, "error": "Rimborso non trovato."}

    data = load_data()
    for item in data["repayments"]:
        if item.get("id") == item_id:
            _apply_update_fields(item, args, "lender")
            save_data(data)
            return {"success": True, "refresh": True}
    return {"success": False, "error": "Rimborso non trovato."}


_AGENT_DISPATCH: dict = {
    "get_summary": agent_fn_get_summary,
    "search_entries": agent_fn_search_entries,
    "get_lender_metrics": agent_fn_get_lender_metrics,
    "verify_lender_amount": agent_fn_verify_lender_amount,
    "add_expense": agent_fn_add_expense,
    "add_loan": agent_fn_add_loan,
    "add_repayment": agent_fn_add_repayment,
    "delete_expense": agent_fn_delete_expense,
    "delete_loan": agent_fn_delete_loan,
    "delete_repayment": agent_fn_delete_repayment,
    "delete_all_entries": agent_fn_delete_all_entries,
    "update_expense": agent_fn_update_expense,
    "update_loan": agent_fn_update_loan,
    "update_repayment": agent_fn_update_repayment,
}


def dispatch_agent_function(name: str, args: dict) -> dict:
    fn = _AGENT_DISPATCH.get(name)
    if fn is None:
        return {"success": False, "error": f"Funzione '{name}' non riconosciuta."}
    try:
        no_arg_fns = {"get_summary"}
        return fn() if name in no_arg_fns else fn(args)
    except Exception as exc:
        return {"success": False, "error": f"Errore esecuzione {name}: {exc}"}


def sanitize_agent_reply_text(reply_text: str, refresh_needed: bool = False) -> str:
    text = sanitize_text(reply_text)
    if text:
        text = re.sub(r"[^.!?\n]*\bID\b[^.!?\n]*[.!?]?\s*", "", text, flags=re.IGNORECASE)

        for raw_name, display_name in CATEGORY_DISPLAY_NAMES.items():
            text = re.sub(rf"\b{re.escape(raw_name)}\b", display_name, text, flags=re.IGNORECASE)

        def _format_iso_date_match(match: re.Match) -> str:
            iso_value = match.group(0)
            try:
                parsed = datetime.fromisoformat(iso_value)
                return parsed.strftime("%d/%m/%Y")
            except ValueError:
                return iso_value

        text = re.sub(r"\b\d{4}-\d{2}-\d{2}\b", _format_iso_date_match, text)
        text = re.sub(r"\n{3,}", "\n\n", text).strip()

    if text:
        return text

    if refresh_needed:
        return "Operazione completata con successo."

    return (
        "Per procedere ho bisogno di un dettaglio in più sulla voce "
        "(ad esempio data, importo, descrizione o prestatore)."
    )


def get_relative_date_hints(user_text: str, base_day: date | None = None) -> list[str]:
    text = sanitize_text(user_text)
    if not text:
        return []

    current_day = base_day or date.today()
    patterns = [
        (r"\boggi\b", "oggi", 0),
        (r"\bieri\b", "ieri", -1),
        (r"\bdomani\b", "domani", 1),
        (r"\bdopodomani\b", "dopodomani", 2),
        (r"\bl[\'’]altro\s+ieri\b", "l'altro ieri", -2),
    ]

    hints: list[str] = []
    for pattern, label, delta_days in patterns:
        if re.search(pattern, text, flags=re.IGNORECASE):
            iso_day = (current_day + timedelta(days=delta_days)).isoformat()
            hints.append(f"- '{label}' = {iso_day}")

    return hints


def call_gemini_agent_chat(history: list[dict]) -> dict:
    """Multi-turn Gemini agent with native Function Calling."""
    history = _normalize_client_history(history, AI_MAX_HISTORY_TURNS)

    api_key = get_gemini_api_key()
    if not api_key:
        return {
            "reply": "GEMINI_API_KEY non configurata. Imposta la variabile d'ambiente e riavvia.",
            "history": history,
            "refresh": False,
        }

    model = get_gemini_model()
    endpoint = (
        f"https://generativelanguage.googleapis.com/v1beta/models/{model}"
        f":generateContent?key={api_key}"
    )
    today_str = date.today().isoformat()
    yesterday_str = (date.today() - timedelta(days=1)).isoformat()
    tomorrow_str = (date.today() + timedelta(days=1)).isoformat()
    day_after_tomorrow_str = (date.today() + timedelta(days=2)).isoformat()
    day_before_yesterday_str = (date.today() - timedelta(days=2)).isoformat()
    system_text = AGENT_SYSTEM_INSTRUCTION.format(
        today=today_str,
        yesterday=yesterday_str,
        tomorrow=tomorrow_str,
        day_after_tomorrow=day_after_tomorrow_str,
        day_before_yesterday=day_before_yesterday_str,
    )

    latest_user_text = ""
    if history:
        last_turn = history[-1]
        if isinstance(last_turn, dict) and last_turn.get("role") == "user":
            latest_user_text = "\n".join(
                sanitize_text((part or {}).get("text", ""))
                for part in (last_turn.get("parts") or [])
                if isinstance(part, dict)
            ).strip()

    relative_date_hints = get_relative_date_hints(latest_user_text, date.today())
    if relative_date_hints:
        system_text += (
            "\n\nContesto rilevato nell'ultimo messaggio utente:\n"
            "Sono presenti riferimenti temporali relativi già sufficienti. "
            "Non richiedere nuovamente la data, converti direttamente in ISO e procedi.\n"
            + "\n".join(relative_date_hints)
        )

    refresh_needed = False

    for _iteration in range(AI_MAX_ITERATIONS):
        body = {
            "system_instruction": {"parts": [{"text": system_text}]},
            "tools": [AGENT_TOOL_DECLARATIONS],
            "contents": history,
            "generationConfig": {
                "maxOutputTokens": AI_MAX_OUTPUT_TOKENS,
            },
        }
        req = urllib.request.Request(
            endpoint,
            data=json.dumps(body).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )

        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
        except urllib.error.HTTPError as exc:
            err_body = exc.read().decode("utf-8", errors="replace")
            msg = ""
            try:
                err_obj = json.loads(err_body).get("error", {})
                msg = sanitize_text(err_obj.get("message", ""))
            except Exception:
                pass
            if exc.code == 429:
                reply = "Quota Gemini esaurita (HTTP 429). Riprova tra poco."
            elif exc.code in {401, 403}:
                reply = "Accesso Gemini negato: controlla la GEMINI_API_KEY."
            else:
                reply = f"Errore Gemini HTTP {exc.code}. {msg or 'Riprova tra poco.'}"
            return {"reply": reply, "history": history, "refresh": False}
        except (urllib.error.URLError, TimeoutError):
            return {
                "reply": "Connessione a Gemini non riuscita (rete/timeout). Riprova.",
                "history": history,
                "refresh": False,
            }

        try:
            gemini_resp = json.loads(raw)
        except Exception:
            return {"reply": "Risposta Gemini non valida (JSON).", "history": history, "refresh": False}

        candidates = gemini_resp.get("candidates", [])
        if not candidates:
            finish = gemini_resp.get("promptFeedback", {}).get("blockReason", "")
            reason_hint = f" (motivo: {finish})" if finish else ""
            return {
                "reply": (
                    "Non ho ricevuto una risposta completa dal modello"
                    f"{reason_hint}. Riprova con una richiesta più specifica."
                ),
                "history": history,
                "refresh": False,
            }

        content = candidates[0].get("content", {})
        parts = content.get("parts", [])

        # Strip [REFRESH] from text parts so history stored client-side is clean.
        clean_parts = [
            {**p, "text": p["text"].replace("[REFRESH]", "").strip()} if "text" in p else p
            for p in parts
        ]
        history = history + [{"role": "model", "parts": clean_parts}]

        function_calls = [p for p in parts if "functionCall" in p]
        if not function_calls:
            text_reply = "\n".join(p.get("text", "") for p in parts if "text" in p).strip()
            refresh_tag = "[REFRESH]" in text_reply
            clean_reply = text_reply.replace("[REFRESH]", "").strip()
            if refresh_tag:
                refresh_needed = True
            final_reply = sanitize_agent_reply_text(clean_reply, refresh_needed=refresh_needed)
            return {
                "reply": final_reply,
                "history": history,
                "refresh": refresh_needed,
            }

        fn_responses = []
        for fc_part in function_calls:
            fc = fc_part["functionCall"]
            fn_name = fc.get("name", "")
            fn_args = fc.get("args") or {}
            result = dispatch_agent_function(fn_name, fn_args)
            if result.get("refresh"):
                refresh_needed = True
            fn_responses.append(
                {"functionResponse": {"name": fn_name, "response": result}}
            )

        history = history + [{"role": "user", "parts": fn_responses}]

    return {
        "reply": "Il processo ha raggiunto il limite massimo di iterazioni. Riprova.",
        "history": history,
        "refresh": refresh_needed,
    }


@app.route("/ai-agent-chat", methods=["POST"])
def ai_agent_chat():
    if not is_authenticated():
        return jsonify({"error": "Non autenticato"}), 401
    payload = request.get_json(silent=True) or {}
    message = sanitize_text(payload.get("message", ""))
    raw_history = payload.get("history")
    history: list[dict] = raw_history if isinstance(raw_history, list) else []
    history = _normalize_client_history(history, AI_MAX_HISTORY_TURNS)

    if not message:
        return jsonify({"reply": "Scrivi un messaggio.", "history": _build_client_safe_history(history, AI_MAX_HISTORY_TURNS), "refresh": False})

    if not _is_same_last_user_turn(history, message):
        history = history + [{"role": "user", "parts": [{"text": message}]}]
    result = call_gemini_agent_chat(history)
    result["history"] = _build_client_safe_history(result.get("history") or history, AI_MAX_HISTORY_TURNS)
    return jsonify(result)

@app.route("/sw.js", methods=["GET"])
def service_worker():
    response = send_from_directory(os.path.join(BASE_DIR, "static"), "sw.js")
    response.headers["Cache-Control"] = "no-cache"
    return response


@app.route("/add-expense", methods=["POST"])
def add_expense():
    try:
        category = sanitize_text(request.form.get("category"))
        if category not in {"acquisto_casa", "ristrutturazione"}:
            raise ValueError("Categoria non valida")

        item = {
            "id": new_id(),
            "date": parse_iso_date(request.form.get("date") or "").isoformat(),
            "description": capitalize_first(sanitize_text(request.form.get("description"))) or "Spesa",
            "amount": parse_amount(request.form.get("amount")),
        }

        if USE_DATABASE:
            with get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO expenses (id, category, operation_date, description, amount)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        (item["id"], category, item["date"], item["description"], item["amount"]),
                    )
                conn.commit()
        else:
            data = load_data()
            data["expenses"][category].append(item)
            save_data(data)
        return redirect(url_for("index"))
    except ValueError as exc:
        return render_index_error(str(exc))


@app.route("/add-loan", methods=["POST"])
def add_loan():
    try:
        item = {
            "id": new_id(),
            "date": parse_iso_date(request.form.get("date") or "").isoformat(),
            "lender": capitalize_first(sanitize_text(request.form.get("lender"))) or "Familiare",
            "note": sanitize_text(request.form.get("note")),
            "amount": parse_amount(request.form.get("amount")),
        }

        if USE_DATABASE:
            with get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO loans (id, operation_date, lender, note, amount)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        (item["id"], item["date"], item["lender"], item["note"], item["amount"]),
                    )
                conn.commit()
        else:
            data = load_data()
            data["loans"].append(item)
            save_data(data)
        return redirect(url_for("index"))
    except ValueError as exc:
        return render_index_error(str(exc))


@app.route("/add-repayment", methods=["POST"])
def add_repayment():
    try:
        item = {
            "id": new_id(),
            "date": parse_iso_date(request.form.get("date") or "").isoformat(),
            "lender": capitalize_first(sanitize_text(request.form.get("lender"))) or "Familiare",
            "amount": parse_amount(request.form.get("amount")),
        }

        if USE_DATABASE:
            with get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO repayments (id, operation_date, lender, amount)
                        VALUES (%s, %s, %s, %s)
                        """,
                        (item["id"], item["date"], item["lender"], item["amount"]),
                    )
                conn.commit()
        else:
            data = load_data()
            data["repayments"].append(item)
            save_data(data)
        return redirect(url_for("index"))
    except ValueError as exc:
        return render_index_error(str(exc))


@app.route("/delete-item", methods=["POST"])
def delete_item():
    section = sanitize_text(request.form.get("section"))
    item_id = sanitize_text(request.form.get("item_id"))
    if not item_id:
        return redirect(url_for("index"))

    if USE_DATABASE:
        deleted = False
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                if section in {"acquisto_casa", "ristrutturazione"}:
                    cur.execute("DELETE FROM expenses WHERE id = %s AND category = %s", (item_id, section))
                elif section == "loans":
                    cur.execute("DELETE FROM loans WHERE id = %s", (item_id,))
                elif section == "repayments":
                    cur.execute("DELETE FROM repayments WHERE id = %s", (item_id,))
                deleted = cur.rowcount > 0
            if deleted:
                conn.commit()
    else:
        data = load_data()
        deleted = False

        if section in {"acquisto_casa", "ristrutturazione"}:
            deleted = remove_by_id(data["expenses"][section], item_id)
        elif section == "loans":
            deleted = remove_by_id(data["loans"], item_id)
        elif section == "repayments":
            deleted = remove_by_id(data["repayments"], item_id)

        if deleted:
            save_data(data)
    return redirect(url_for("index"))


@app.route("/edit-item", methods=["POST"])
def edit_item():
    section = sanitize_text(request.form.get("section"))
    item_id = sanitize_text(request.form.get("item_id"))
    label = capitalize_first(sanitize_text(request.form.get("label")))
    loan_note = sanitize_text(request.form.get("note"))
    date_raw = request.form.get("date") or ""

    try:
        amount = parse_amount(request.form.get("amount"))
        operation_date = parse_iso_date(date_raw).isoformat()
    except ValueError as exc:
        return render_index_error(str(exc))

    if not item_id or not label:
        return render_index_error("Compila nome e importo per modificare la voce")

    if USE_DATABASE:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                if section in {"acquisto_casa", "ristrutturazione"}:
                    cur.execute(
                        """
                        UPDATE expenses
                        SET description = %s, amount = %s, operation_date = %s
                        WHERE id = %s AND category = %s
                        """,
                        (label, amount, operation_date, item_id, section),
                    )
                elif section == "loans":
                    cur.execute(
                        """
                        UPDATE loans
                        SET lender = %s, note = %s, amount = %s, operation_date = %s
                        WHERE id = %s
                        """,
                        (label, loan_note, amount, operation_date, item_id),
                    )
                elif section == "repayments":
                    cur.execute(
                        """
                        UPDATE repayments
                        SET lender = %s, amount = %s, operation_date = %s
                        WHERE id = %s
                        """,
                        (label, amount, operation_date, item_id),
                    )
                else:
                    return redirect(url_for("index"))

                if cur.rowcount > 0:
                    conn.commit()
    else:
        data = load_data()
        updated = False
        if section in {"acquisto_casa", "ristrutturazione"}:
            updated = update_local_item(
                data["expenses"][section],
                item_id,
                "description",
                label,
                amount,
                operation_date,
            )
        elif section == "loans":
            updated = update_local_item(data["loans"], item_id, "lender", label, amount, operation_date)
            if updated:
                for item in data["loans"]:
                    if item.get("id") == item_id:
                        item["note"] = loan_note
                        break
        elif section == "repayments":
            updated = update_local_item(data["repayments"], item_id, "lender", label, amount, operation_date)

        if updated:
            save_data(data)

    return redirect(url_for("index"))


def build_excel_workbook(data: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    summary = build_summary(data)
    amount_format = "#,##0.00"

    def apply_amount_style(sheet, column_index: int, start_row: int = 2) -> None:
        for row_index in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            if isinstance(cell.value, (int, float)):
                cell.number_format = amount_format
                cell.alignment = Alignment(horizontal="right")

    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo"

    ws.append(["Voce", "Valore"])
    ws.append(["Spese acquisto casa", summary["acquisto_total"]])
    ws.append(["Spese ristrutturazione", summary["ristr_total"]])
    ws.append(["Spese totali", summary["spese_total"]])
    ws.append(["Prestiti ricevuti", summary["loans_total"]])
    ws.append(["Rimborsi effettuati", summary["repayments_total"]])
    ws.append(["Debito residuo prestiti", summary["debito_residuo"]])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col in ("A", "B"):
        ws.column_dimensions[col].width = 28
    apply_amount_style(ws, column_index=2)

    ws_lender = wb.create_sheet("Saldo per prestatore")
    ws_lender.append(["Prestatore", "Saldo residuo"])
    for row in summary["lender_balance"]:
        ws_lender.append([row["lender"], row["balance"]])
    for cell in ws_lender[1]:
        cell.font = Font(bold=True)
    ws_lender.column_dimensions["A"].width = 24
    ws_lender.column_dimensions["B"].width = 20
    apply_amount_style(ws_lender, column_index=2)

    def append_sheet(title: str, entries: list[dict], include_lender: bool = False, include_note: bool = False):
        sheet = wb.create_sheet(title)
        if include_lender and include_note:
            sheet.append(["Data", "Prestatore", "Note", "Importo"])
        elif include_lender:
            sheet.append(["Data", "Prestatore", "Importo"])
        else:
            sheet.append(["Data", "Descrizione", "Importo"])

        for item in sort_entries(entries):
            if include_lender and include_note:
                sheet.append([item.get("date", ""), item.get("lender", ""), item.get("note", ""), item.get("amount", 0.0)])
            elif include_lender:
                sheet.append([item.get("date", ""), item.get("lender", ""), item.get("amount", 0.0)])
            else:
                sheet.append([item.get("date", ""), item.get("description", ""), item.get("amount", 0.0)])

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 24
        if include_lender and include_note:
            sheet.column_dimensions["C"].width = 36
            sheet.column_dimensions["D"].width = 14
            apply_amount_style(sheet, column_index=4)
        else:
            sheet.column_dimensions["C"].width = 14
            apply_amount_style(sheet, column_index=3)

    append_sheet("Acquisto casa", data["expenses"]["acquisto_casa"])
    append_sheet("Ristrutturazione", data["expenses"]["ristrutturazione"])
    append_sheet("Prestiti ricevuti", data["loans"], include_lender=True, include_note=True)
    append_sheet("Rimborsi", data["repayments"], include_lender=True)

    return wb


@app.route("/export/excel", methods=["GET"])
def export_excel():
    try:
        workbook = build_excel_workbook(load_data())
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = f"home13_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        return render_index_error(f"Errore export Excel: {exc}")


@app.route("/export/pdf", methods=["GET"])
def export_pdf():
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except Exception as exc:
        return render_index_error(f"Errore export PDF: {exc}")

    data = load_data()
    summary = build_summary(data)

    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            self._saved_page_states.append(dict(self.__dict__))
            page_count = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                draw_page_chrome(self, self._pageNumber, page_count)
                canvas.Canvas.showPage(self)
            canvas.Canvas.save(self)

    buffer = io.BytesIO()
    pdf = NumberedCanvas(buffer, pagesize=A4)
    width, height = A4
    left = 34
    right = width - 34
    content_top = height - 72
    content_bottom = 42

    generated_on = format_date_it(date.today().isoformat())
    palette = {
        "brand": colors.HexColor("#0f4f49"),
        "brand_dark": colors.HexColor("#123f3a"),
        "ink_soft": colors.HexColor("#5d746e"),
        "line": colors.HexColor("#d7e4e0"),
        "line_soft": colors.HexColor("#e3ece9"),
        "header_bg": colors.HexColor("#eef5f3"),
        "row_alt": colors.HexColor("#f9fcfb"),
        "card_bg": colors.HexColor("#f4f9f8"),
    }

    def draw_page_chrome(cnv, page_number: int, page_count: int) -> None:
        cnv.saveState()
        cnv.setStrokeColor(palette["line"])
        cnv.setLineWidth(0.8)
        cnv.line(left, height - 32, right, height - 32)
        cnv.line(left, 24, right, 24)

        cnv.setFillColor(palette["brand"])
        cnv.setFont("Helvetica-Bold", 9)
        cnv.drawString(left, height - 25, "Home13 - Report Spese")

        cnv.setFillColor(palette["ink_soft"])
        cnv.setFont("Helvetica", 8)
        cnv.drawRightString(right, height - 25, f"Generato il {generated_on}")
        cnv.drawString(left, 14, "Documento riservato")
        cnv.drawRightString(right, 14, f"Pag. {page_number}/{page_count}")
        cnv.restoreState()

    def new_page(current_y: float) -> float:
        if current_y < content_bottom:
            pdf.showPage()
            return content_top
        return current_y

    def draw_title(text: str, current_y: float) -> float:
        current_y = new_page(current_y)
        pdf.setFillColor(palette["brand_dark"])
        pdf.setFont("Helvetica-Bold", 18)
        pdf.drawString(left, current_y, text)
        pdf.setStrokeColor(palette["line"])
        pdf.setLineWidth(1)
        pdf.line(left, current_y - 10, right, current_y - 10)
        return current_y - 28

    def draw_section_header(text: str, current_y: float) -> float:
        current_y = new_page(current_y)
        pdf.setFillColor(palette["brand_dark"])
        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawString(left, current_y, text)
        pdf.setStrokeColor(palette["line_soft"])
        pdf.setLineWidth(0.8)
        pdf.line(left, current_y - 3, right, current_y - 3)
        return current_y - 20

    def draw_metric_card(x: float, top_y: float, card_w: float, label: str, value: str) -> None:
        card_h = 56
        pdf.setFillColor(palette["card_bg"])
        pdf.setStrokeColor(palette["line"])
        pdf.roundRect(x, top_y - card_h, card_w, card_h, 8, fill=1, stroke=1)

        pdf.setFillColor(palette["ink_soft"])
        pdf.setFont("Helvetica", 8)
        pdf.drawString(x + 10, top_y - 16, label)

        pdf.setFillColor(palette["brand_dark"])
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(x + 10, top_y - 37, value)

    def draw_summary_cards(current_y: float) -> float:
        card_gap = 10
        card_w = (right - left - card_gap) / 2
        top = current_y

        draw_metric_card(left, top, card_w, "Spese totali", f"EUR {format_euro(summary['spese_total'])}")
        draw_metric_card(left + card_w + card_gap, top, card_w, "Debito residuo", f"EUR {format_euro(summary['debito_residuo'])}")

        top2 = top - 66
        draw_metric_card(left, top2, card_w, "Prestiti ricevuti", f"EUR {format_euro(summary['loans_total'])}")
        draw_metric_card(left + card_w + card_gap, top2, card_w, "Rimborsi effettuati", f"EUR {format_euro(summary['repayments_total'])}")

        return top2 - 74

    def draw_category_bars(current_y: float) -> float:
        current_y = draw_section_header("Sintesi categorie", current_y)
        bar_left = left
        bar_right = right
        bar_width = bar_right - bar_left

        series = [
            ("Acquisto casa", summary["acquisto_total"], colors.HexColor("#4c89e8")),
            ("Ristrutturazione", summary["ristr_total"], colors.HexColor("#ef5c58")),
            ("Prestiti ricevuti", summary["loans_total"], colors.HexColor("#21a77f")),
            ("Rimborsi", summary["repayments_total"], colors.HexColor("#f08c4a")),
        ]
        max_value = max((row[1] for row in series), default=1.0) or 1.0

        y_pos = current_y
        for label, value, color in series:
            y_pos = new_page(y_pos)
            pdf.setFillColor(palette["ink_soft"])
            pdf.setFont("Helvetica", 9)
            pdf.drawString(bar_left, y_pos, label)
            pdf.drawRightString(right, y_pos, f"EUR {format_euro(value)}")

            y_pos -= 8
            pdf.setFillColor(palette["header_bg"])
            pdf.roundRect(bar_left, y_pos - 8, bar_width, 8, 3, fill=1, stroke=0)
            scaled_w = max((value / max_value) * bar_width, 0.0)
            pdf.setFillColor(color)
            pdf.roundRect(bar_left, y_pos - 8, scaled_w, 8, 3, fill=1, stroke=0)
            y_pos -= 18

        return y_pos - 12

    def draw_table(title: str, headers: list[str], rows: list[list[str]], col_widths: list[float], current_y: float) -> float:
        section_gap_to_table = 14
        block_gap_after_table = 24
        row_h = 20
        title_block_h = 13

        def ensure_table_block_fits(y_top: float) -> float:
            # Keep section title and table header together, plus at least one row (or empty state line).
            min_body_h = row_h if rows else 16
            required_h = title_block_h + section_gap_to_table + row_h + min_body_h
            if y_top - required_h < content_bottom:
                pdf.showPage()
                return content_top
            return y_top

        def draw_table_title(y_top: float) -> float:
            pdf.setFillColor(palette["brand_dark"])
            pdf.setFont("Helvetica-Bold", 13)
            pdf.drawString(left, y_top, title)
            pdf.setStrokeColor(palette["line_soft"])
            pdf.setLineWidth(0.8)
            pdf.line(left, y_top - 3, right, y_top - 3)
            return y_top - section_gap_to_table

        def draw_header_and_frame(y_top: float) -> float:
            pdf.setFillColor(palette["header_bg"])
            pdf.rect(left, y_top - row_h, sum(col_widths), row_h, fill=1, stroke=0)

            x = left
            pdf.setFillColor(palette["brand_dark"])
            pdf.setFont("Helvetica-Bold", 9)
            for idx, header in enumerate(headers):
                pdf.drawString(x + 5, y_top - 12, header)
                x += col_widths[idx]

            pdf.setStrokeColor(palette["line"])
            pdf.setLineWidth(0.6)
            pdf.rect(left, y_top - row_h, sum(col_widths), row_h, fill=0, stroke=1)
            return y_top - row_h

        current_y = ensure_table_block_fits(current_y)
        current_y = draw_table_title(current_y)

        if not rows:
            pdf.setFillColor(palette["ink_soft"])
            pdf.setFont("Helvetica", 9)
            pdf.drawString(left, current_y, "Nessun dato disponibile")
            return current_y - block_gap_after_table

        current_y = draw_header_and_frame(current_y)

        for idx, row in enumerate(rows):
            if current_y - row_h < content_bottom:
                pdf.showPage()
                current_y = content_top
                current_y = draw_header_and_frame(current_y)

            if idx % 2 == 0:
                pdf.setFillColor(palette["row_alt"])
                pdf.rect(left, current_y - row_h, sum(col_widths), row_h, fill=1, stroke=0)

            x = left
            pdf.setFillColor(palette["brand_dark"])
            pdf.setFont("Helvetica", 9)
            for col_idx, cell in enumerate(row):
                text = str(cell)
                if col_idx == len(row) - 1:
                    pdf.drawRightString(x + col_widths[col_idx] - 5, current_y - 12, text)
                else:
                    pdf.drawString(x + 5, current_y - 12, text)
                x += col_widths[col_idx]

            pdf.setStrokeColor(palette["line_soft"])
            pdf.setLineWidth(0.5)
            pdf.line(left, current_y - row_h, left + sum(col_widths), current_y - row_h)
            current_y -= row_h

        return current_y - block_gap_after_table

    y = content_top
    y = draw_title("Report Home13", y)
    y = draw_summary_cards(y)
    y = draw_category_bars(y)

    acquisto_rows = [
        [format_date_it(item.get("date", "")), item.get("description", ""), f"EUR {format_euro(item.get('amount', 0.0))}"]
        for item in sort_entries(data["expenses"]["acquisto_casa"])
    ]
    ristr_rows = [
        [format_date_it(item.get("date", "")), item.get("description", ""), f"EUR {format_euro(item.get('amount', 0.0))}"]
        for item in sort_entries(data["expenses"]["ristrutturazione"])
    ]
    prestiti_rows = [
        [
            format_date_it(item.get("date", "")),
            item.get("lender", ""),
            item.get("note", ""),
            f"EUR {format_euro(item.get('amount', 0.0))}",
        ]
        for item in sort_entries(data["loans"])
    ]
    rimborsi_rows = [
        [format_date_it(item.get("date", "")), item.get("lender", ""), f"EUR {format_euro(item.get('amount', 0.0))}"]
        for item in sort_entries(data["repayments"])
    ]
    saldo_rows = [[row["lender"], f"EUR {format_euro(row['balance'])}"] for row in summary["lender_balance"]]

    y = draw_table("Acquisto casa", ["Data", "Voce", "Importo"], acquisto_rows, [110, 300, 110], y)
    y = draw_table("Ristrutturazione", ["Data", "Voce", "Importo"], ristr_rows, [110, 300, 110], y)
    y = draw_table("Prestiti ricevuti", ["Data", "Prestatore", "Note", "Importo"], prestiti_rows, [85, 135, 190, 110], y)
    y = draw_table("Rimborsi", ["Data", "Prestatore", "Importo"], rimborsi_rows, [110, 300, 110], y)
    draw_table("Saldo per prestatore", ["Prestatore", "Saldo residuo"], saldo_rows, [380, 140], y)

    pdf.save()
    buffer.seek(0)
    filename = f"home13_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


if __name__ == "__main__":
    raw_port = (os.environ.get("PORT") or "").strip()
    port = int(raw_port) if raw_port.isdigit() else 5000
    debug_mode = (os.environ.get("FLASK_DEBUG") or "").strip() == "1"
    app.run(host="0.0.0.0", port=port, debug=debug_mode)
